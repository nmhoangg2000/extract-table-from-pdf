from openpyxl import load_workbook
import re
# Load the workbook and select the active sheet
wb = load_workbook(r'E:\mcc\extract_table\updated_merged_tables_lattice1.xlsx')
ws = wb['Table 2']

# Iterate through columns B and C to fill blank cells with the value of the first non-empty cell above them
for col in ['B', 'C']:  # Specify the columns you want to fill
    fill_value = None
    for row in range(2, ws.max_row + 1):  # Start at row 2 to skip the header
        cell = ws[col + str(row)]
        if cell.value not in [None, '']:  # If the cell is not empty
            fill_value = cell.value  # Update the fill value
        else:  # If the cell is empty
            cell.value = fill_value  # Set it to the fill value

# Remove any rows that are completely blank, starting from the bottom to avoid skipping rows after deletion
for row in range(ws.max_row, 1, -1):
    if all(ws.cell(row=row, column=col).value in [None, ''] for col in range(1, ws.max_column + 1)):
        ws.delete_rows(row)






# Save the workbook
wb.save(r'E:\mcc\extract_table\perfect.xlsx')
filled_and_cleaned_path = 'perfect.xlsx'
filled_and_cleaned_path