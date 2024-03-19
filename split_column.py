from openpyxl import load_workbook
import re

# Load the workbook and select sheet 2
wb = load_workbook('E:\mcc\extract_table\merged_tables_lattice1.xlsx')  # Make sure to put the correct file path here
ws = wb['Table 2']  # Change 'Sheet2' to the correct sheet name if different

# Define the regular expressions
# bank_regex = re.compile(r"Ngân hàng[\w\s]*")
# bank_regex = re.compile(r"Ngân hàng[\w\s]+|Chi nhánh[\w\s]+|Chinhánh[\w\s]+")
bank_regex = re.compile(r"(\d{8}-Ngân hàng[\w\s]+(?:- Chi| - Chi| - Chinhánh)[\w\s]+)")

date_regex = re.compile(r"\d{2}/\d{2}/\d{4}")
debt_regex = re.compile(r"Dư nợ[\w\s]*")

# Insert new columns for the extracted information
ws.insert_cols(2, 3)
# Set the headers for the new columns
ws['B1'] = 'Tên tổ chức'
ws['C1'] = 'Ngày báo cáo'
ws['D1'] = 'Loại dư nợ'
# Iterate over the rows in what was originally column B and is now column E (due to the insertions)
for row in ws.iter_rows(min_col=5, max_col=5, min_row=2):  # min_row=2 to skip the header
    cell = row[0]
    cell_value = cell.value or ""
    
    # Extract information
    bank_match = bank_regex.search(cell_value)
    date_match = date_regex.search(cell_value)
    debt_match = debt_regex.search(cell_value)

    # Place the extracted values in the new columns
    ws.cell(row=cell.row, column=2).value = bank_match.group(0) if bank_match else None
    ws.cell(row=cell.row, column=3).value = date_match.group(0) if date_match else None
    ws.cell(row=cell.row, column=4).value = debt_match.group(0) if debt_match else None

# Delete the original data column, which has been shifted to column E after the insertions
ws.delete_cols(5)

# Save the workbook with the new structure
wb.save('updated_merged_tables_lattice1.xlsx')  # Specify your desired new file path or overwrite the original
