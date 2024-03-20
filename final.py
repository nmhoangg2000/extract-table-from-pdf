from openpyxl import load_workbook

# Load the workbook and specify the path to your Excel file
wb = load_workbook(r'E:\mcc\extract_table\perfect.xlsx')

# Access the first sheet to get the values from the first row
first_sheet = wb[wb.sheetnames[0]]
# header_values = [cell.value for cell in first_sheet[2]]  # Assuming header is in the first row

# Access the second sheet where you want to add a new column with these values
second_sheet = wb[wb.sheetnames[1]]


# Insert a new column in sheet 2 at the beginning
# second_sheet.insert_cols(2)


# Identify all rows with a blank cell in column 'E' in the second sheet and delete them
rows_to_delete = [row.row for row in second_sheet['F'] if row.value in [None, ""]]
for row_idx in reversed(sorted(rows_to_delete)):
    second_sheet.delete_rows(row_idx)


# # Set the values of the first column in sheet 2 to the header values from sheet 1
# for index, value in enumerate(header_values, start=1):  # start=1 to fill from the first row
#     second_sheet.cell(row=index, column=2).value = value

rows_to_delete = []
for row in second_sheet.iter_rows(min_row=1, max_col=4, values_only=True):
    # Check if the cell in column 'C' is empty (None or empty string)
    if row[2] in [None, ""]:
        rows_to_delete.append(row[0])

#Iterate through each sheet in the workbook
sheet_names = wb.sheetnames
for name in sheet_names:
    sheet = wb[name]

    # Iterate over all cells in all rows
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value and isinstance(cell.value, str):
                # Replace line breaks with spaces
                cell.value = cell.value.replace('_x000D_', ' ')



# Reverse the rows to delete to avoid index shifting during deletion
rows_to_delete.reverse()


# # Function to check if a row is blank
# def is_row_blank(row):
#     return all(cell.value is None for cell in row)

# # Iterate over all sheets in the workbook
# for sheet_name in wb.sheetnames:
#     sheet = wb[sheet_name]
#     blank_row_indexes = []
    
#     # Identify blank rows
#     for row in sheet.iter_rows():
#         if is_row_blank(row):
#             blank_row_indexes.append(row[0].row)
    
#     # Delete blank rows from the bottom up
#     for row_index in reversed(blank_row_indexes):
#         sheet.delete_rows(row_index)

# Values to be used as headers and their respective first row data
header_and_first_row_values = [
    (wb[wb.sheetnames[0]]['B1'].value, wb[wb.sheetnames[0]]['C1'].value),  # Pair for the first new column
    (wb[wb.sheetnames[0]]['B2'].value, wb[wb.sheetnames[0]]['C2'].value)   # Pair for the second new column
]

# Iterate over all sheets except the first one
for sheet_name in wb.sheetnames[1:]:
    sheet = wb[sheet_name]
    
    # Insert two new columns at the beginning
    sheet.insert_cols(2)
    sheet.insert_cols(2)
    
    # Set headers and first row values for the new columns
    for i, (header, first_row_value) in enumerate(header_and_first_row_values, start=2):
        sheet.cell(row=1, column=i, value=header)
        sheet.cell(row=2, column=i, value=first_row_value)

         # Fill any remaining blank cells in these columns with the respective values
        max_row = sheet.max_row
        for row in range(3, max_row + 1):  # Start filling from row 3
            if sheet.cell(row=row, column=i).value is None:
                sheet.cell(row=row, column=i, value=first_row_value)


# Iterate over all sheets in the workbook
for sheet_name in wb.sheetnames:
    sheet = wb[sheet_name]
    # Delete the first column (Column A)
    sheet.delete_cols(1)


from openpyxl.utils import get_column_letter

# Function to combine cell values, avoiding internal duplicates
def combine_unique_values(upper_value, lower_value):
    combined = f"{upper_value}, {lower_value}".strip()  # Initial combination
    unique_values = set(combined.split(", "))  # Split by comma and convert to set for uniqueness
    return ', '.join(unique_values)  # Re-join unique values

# Iterate over all sheets in the workbook
for sheet_name in wb.sheetnames:
    sheet = wb[sheet_name]

    # Find the 'STT' column index
    stt_col_index = None
    for col in sheet.iter_cols(max_row=1):
        for cell in col:
            if cell.value == 'STT':
                stt_col_index = cell.col_idx
                break
        if stt_col_index:
            break

    # If 'STT' column is not found, skip this sheet
    if not stt_col_index:
        continue  # Proceed to the next sheet

    # Check adjacent cells for duplicates in the 'STT' column
    row = 2
    while row < sheet.max_row:
        current_stt = sheet.cell(row=row, column=stt_col_index).value
        next_stt = sheet.cell(row=row+1, column=stt_col_index).value
        
        if current_stt == next_stt:
            # Loop over all columns to combine values, avoiding internal duplicates
            for col in range(1, sheet.max_column + 1):
                upper_cell = sheet.cell(row=row, column=col)
                lower_cell = sheet.cell(row=row+1, column=col)
                # Skip combining 'STT' column values
                if col != stt_col_index:
                    upper_cell.value = combine_unique_values(upper_cell.value, lower_cell.value)
            # Delete the next row after combining values
            sheet.delete_rows(row+1)
        else:
            row += 1  # Increment row index if not deleting


# Save the updated workbook to a new file
wb.save('final2.xlsx')
